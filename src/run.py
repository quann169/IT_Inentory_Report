'''
Created on Nov 11, 2019

@author: phuongtruong
'''
#!/usr/bin/env python
# -*- encoding: utf-8 -*-
import glob
from os import path
import os
import os.path
from shutil import copy
from svi.config import list_special_software
from svi.controller import Parser
from svi.view.Enum import DB_FILE, DB_TAG, DB_COLUMN, DB_ROW
import time

from bs4 import BeautifulSoup
from openpyxl import load_workbook
import openpyxl
from pathlib import Path


#Get current dir of user
current_dir = os.getcwd()
print (current_dir)
# iuput data
list_file_htm = Path(current_dir).glob(DB_FILE.REPORT_EXT)
# print (list_file_htm)
list_file_xlsx = Path(current_dir).glob(DB_FILE.REPORT_EXT_OUT)
#output data
path_out_file = Path(current_dir).glob(DB_FILE.out_dict)

out_list_file_xlsx = Path(current_dir).glob(DB_FILE.out_file)


# Check if file exists and copy file to new file 
for in_xlsx_file in list_file_xlsx:
    print in_xlsx_file
    if os.path.isfile(str(in_xlsx_file)) is True:
        if os.path.basename(str(in_xlsx_file)) == 'SVIHCM_PC_LIST_Summary.xlsx':    
            copy(str(in_xlsx_file), current_dir)
    else:    
            print('Excel input file is not exist')
            
#Clear data in column   
for out_xlsx_file in out_list_file_xlsx:
    print out_xlsx_file
    if os.path.isfile(str(out_xlsx_file)) is True:
        if os.path.basename(str(out_xlsx_file)) == 'SVIHCM_PC_LIST_Summary.xlsx':    
            wbk = openpyxl.load_workbook(str(out_xlsx_file))
            wks = wbk['Summary']    
            list_all_columns = [DB_COLUMN.PC_NAME, DB_COLUMN.MODEL, DB_COLUMN.SERIAL_NO, DB_COLUMN.MAC_ADDRESS, DB_COLUMN.CPU, DB_COLUMN.RAM, DB_COLUMN.HDD, DB_COLUMN.SPECIAL_SOFTWARE, DB_COLUMN.LICENSES, DB_COLUMN.ANTI_VIRUS, DB_COLUMN.BATTERY, DB_COLUMN.USER_NAME, DB_COLUMN.USERS_GROUP, DB_COLUMN.START_UP]      
            for i in range(DB_ROW.START_ROW, DB_ROW.END_ROW):
                for column_name in list_all_columns:
                    wks.cell(row = i, column = column_name).value = None      
            wbk.save(str(out_xlsx_file))
            wbk.close   

for file in list_file_htm:
    start_time = time.clock()

    print "Process file: %s" %(str(file))
    # Open and read file 
    f = open(str(file))
    data = f.read()        
         
#     try:        
    # open file excel
    wbk = openpyxl.load_workbook(str(out_xlsx_file))
        #open sheet necessary
    wks = wbk['Summary']             
                
    #Get data with BeautyfulSoup function
    soup = BeautifulSoup(data, 'lxml')
    # find data in file
    all_table = soup.find_all(Parser.parsing_TABLE.table)

    index_count = 0
        #GET DATA
    for table_data in all_table:
        # GET DATA IN SUMMARY
        if table_data.text == 'Summary':
            all_tr_summary_table = all_table[index_count + 1].find_all(DB_TAG.TR)               
            for i in all_tr_summary_table[2:]:
                all_td = i.find_all(DB_TAG.TD)
                
                if len(all_td) > 4:
                    str_all_td = (all_td[3].text).encode('ascii', 'ignore')
                            
                    # Get PC NAME
                    try:                        
                        if str_all_td == 'Computer Name':
                            print ('PC NAME: ', all_td[4].text)
                            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.PC_NAME).value = all_td[4].text 
                            wbk.save(str(out_xlsx_file))                             
                    except:
                        print ('PC Name is not exist')    
                                                             
                    # Get User Name
                    try:
                        if str_all_td == 'User Name':
                            print ('User Name : ', all_td[4].text)
                            wks.cell( row = DB_ROW.START_ROW, column = DB_COLUMN.USER_NAME).value = all_td[4].text
                            wbk.save(str(out_xlsx_file))
                    except:
                        print ("User Name is not exist")   
                                  
                            # Get CPU Type
                    try:                                
                        if str_all_td == 'CPU Type':
                            print ('CPU Type: ', all_td[4].text)
                            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.CPU).value = all_td[4].text
                            wbk.save(str(out_xlsx_file))
                    except:
                        print('CPU is not exist')
                                
                            #Get motherboard Name/Model
                    try:
                        if str_all_td == 'Motherboard Name':
                            print ('Motherboard Name: ', all_td[4].text)
                            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.MODEL).value = all_td[4].text
                            wbk.save(str(out_xlsx_file))
                    except:
                        print('MODEL is not exist')
                                
                    # get system Memory/RAM
                    try:                                
                        if str_all_td == 'System Memory':
                            print ('System Memory: ', all_td[4].text)
                            str_RAM = all_td[4].text    
                            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.RAM).value = all_td[4].text
                            wbk.save(str(out_xlsx_file))
    #                                 print(type(wks.cell(row = m + 1, column = DB_COLUMN.RAM).value))
                    except:
                        print('RAM is not exist')    
                                       
                        #Get Disk Driver/HDD
                    try:
                        if str_all_td == 'Disk Drive':
                            print ('Disk Drive: ', all_td[4].text)
                            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.HDD).value = all_td[4].text
                            wbk.save(str(out_xlsx_file))
                    except:
                        print('HDD is not exist')
                                                                           

                        #Get serial No.
                    try:
                        if str_all_td == 'DMI System Serial Number':
                            print ('Serial No.: ', all_td[4].text)
                            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.SERIAL_NO).value = all_td[4].text
                            wbk.save(str(out_xlsx_file))
                    except:
                        print ('Serial No. is not exist')              
          
        #Get RAM
        elif table_data.text == "DMI":
            all_tr_RAM = all_table[index_count + 1].find_all(DB_TAG.TR) 
            for j in all_tr_RAM:
                all_td = j.find_all(DB_TAG.TD)
                if len(all_td) > 3:
        
                    str_all_td = (all_td[3].text).encode('ascii', 'ignore')
        
                    str_all_td_1 = (all_td[4].text).encode('ascii', 'ignore')

                    # check Type of RAM
                    if str_all_td == 'Type' and str_all_td_1 =='DDR3\n':
                        print('RAM DATA', str_all_td_1)    
                        wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.RAM).value = wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.RAM).value +'-' + str_all_td_1
                        wbk.save(str(out_xlsx_file))
                        break
                    elif str_all_td == 'Type' and str_all_td_1 == 'DDR4\n':
                        print('ram data', str_all_td_1)
                        wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.RAM).value = wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.RAM).value +'-' + str_all_td_1
                        wbk.save(str(out_xlsx_file))
                        break
                       
        #SPECIAL SOFTWARE
        #Get included software                           
        elif table_data.text == 'Installed Programs':
            all_tr_special_software_table = all_table[index_count + 1].find_all(DB_TAG.TR)
            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.SPECIAL_SOFTWARE).value = ''
            for j in all_tr_special_software_table[2:]:
                all_td = j.find_all(DB_TAG.TD)
                        
                #transfer all_td from unicode to string
                all_td_str = (all_td[2].text).encode('ascii', 'ignore')
                        
                all_td_str_split = all_td_str.split(" ")
                is_legal_software = False
                for split_name in all_td_str_split:
                    if split_name.lower() in [x.lower() for x in list_special_software]:
                        is_legal_software = True
                                
                if not is_legal_software:
                    special_software = all_td_str + ' - ' + all_td[6].text
                    print ('Special software:', all_td_str + ' - ' + all_td[6].text)
                    wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.SPECIAL_SOFTWARE).value = unicode(wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.SPECIAL_SOFTWARE).value) + special_software +'\n'
                    wbk.save(str(out_xlsx_file))
                           
        # POWER MANAGEMENT
        # Get battery status   
        elif table_data.text == 'Power Management':
            all_tr_battery = all_table[index_count + 1].find_all(DB_TAG.TR)
            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.BATTERY).value = ''
            for j in all_tr_battery[2:]:
                all_td= j.find_all(DB_TAG.TD)
                if len(all_td) >3:                    
                    str_all_td = (all_td[3].text).encode('ascii', 'ignore') 
                    str_all_td_1 = (all_td[4].text).encode('ascii', 'ignore') 
                            
                    #check Battery Status     
                    if str_all_td  == 'Battery Status':
                        if str_all_td_1 == 'No Battery\n':
                            print (str_all_td_1)
                            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.BATTERY).value = str_all_td_1
                            wbk.save(str(out_xlsx_file))
                        else:
                            print (str_all_td_1)
                            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.BATTERY).value = unicode(wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.BATTERY).value) + str_all_td_1
                            wbk.save(str(out_xlsx_file))
                                    
                    # Check Wear Level
                    if str_all_td == 'Wear Level':            
                        print (str_all_td_1)
                        wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.BATTERY).value = str_all_td_1+'- '+ unicode(wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.BATTERY).value)
                        wbk.save(str(out_xlsx_file))

        #LICENSES
        # Get all licenses
        elif table_data.text == 'Licenses':
            all_tr_license = all_table[index_count + 1].find_all(DB_TAG.TR)
            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.LICENSES).value = ''
            for j in all_tr_license[2:]:
                all_td = j.find_all(DB_TAG.TD)
                licenses = all_td[2].text + '- ' + all_td[3].text
                print('Licenses: ', all_td[2].text + '- ' + all_td[3].text)
                wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.LICENSES).value = unicode(wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.LICENSES).value) + licenses
                wbk.save(str(out_xlsx_file))
    
                              
        #ANTI-VIRUS
        # Get all startup
        elif table_data.text == 'Anti-Virus':
            all_tr_antivirus = all_table[index_count + 1].find_all(DB_TAG.TR)
            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.ANTI_VIRUS).value = ''
            for j in all_tr_antivirus[2:]:
                all_td = j.find_all(DB_TAG.TD)
                anti_virus = all_td[2].text + '- ' + all_td[3].text + '- ' + all_td[4].text + '- ' + all_td[5].text
                print('Anti-Virus: ', all_td[2].text + '- ' + all_td[3].text + '- ' + all_td[4].text + '- ' + all_td[5].text) 
                wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.ANTI_VIRUS).value =unicode(wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.ANTI_VIRUS).value) + anti_virus
                wbk.save(str(out_xlsx_file))
     
        #Auto Start
        # get all startup       
        elif table_data.text == 'Auto Start':
            all_tr_startup = all_table[index_count + 1].find_all(DB_TAG.TR)
            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.START_UP).value = ''
            for j in all_tr_startup[2:]:
                all_td = j.find_all(DB_TAG.TD)
                start_up = all_td[2].text + '- ' + all_td[4].text
                print ('Startup: ', all_td[2].text + '- ' + all_td[4].text)
                wks.cell(row = DB_ROW.START_ROW,column = DB_COLUMN.START_UP).value =unicode(wks.cell(row =DB_ROW.START_ROW, column = DB_COLUMN.START_UP).value) + start_up
                wbk.save(str(out_xlsx_file))
           
        # AIDA64 EXtreme
        #get generator
        elif table_data.text == 'AIDA64 Extreme':
            all_tr_generator = all_table[index_count + 1].find_all(DB_TAG.TR)
            for j in all_tr_generator:
                all_td = j.find_all(DB_TAG.TD)
                if len(all_td) > 2:
                    str_all_td = (all_td[2].text).encode('ascii', 'ignore')
                    str_all_td_name = (all_td[3].text).encode('ascii', 'ignore')
                    if str_all_td == "Generator":
                        stri = all_td[3].text
                        print(str_all_td_name)

        # USERS                
        # get member of groups           
        elif table_data.text == 'Users':
            all_tr_user = all_table[index_count + 1].find_all(DB_TAG.TR)
            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.USERS_GROUP).value =''
            index = 0
            for j in all_tr_user:

                all_td = j.find_all(DB_TAG.TD)

                if len(all_td) > 3:
                    str_all_td = (all_td[3].text).encode('ascii', 'ignore')
                    str_all_td_1 = (all_td[4].text).encode('ascii', 'ignore')

                        #compare with generator name
                    if str_all_td == 'User Name' and str_all_td_1 == stri:
                        all_td = all_tr_user[index+2].find_all(DB_TAG.TD)
                        print ('Member of groups:', all_td[4].text)
                        wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.USERS_GROUP).value = all_td[4].text
                        wbk.save(str(out_xlsx_file))
                index+=1

        #Windows Network
        #get MAC Address (Wifi & Ethernet)  
        elif table_data.text == 'Windows Network':
            all_tr_wifi = all_table[index_count + 1].find_all(DB_TAG.TR)
            index = 0
            wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.MAC_ADDRESS).value = ''
                    
            for j in all_tr_wifi:
                all_td = j.find_all(DB_TAG.TD)  
                                              
                if len(all_td) > 3:
                    str_all_td = (all_td[3].text).encode('ascii', 'ignore')
                    str_all_td_1 = (all_td[4].text).encode('ascii', 'ignore')
                    if str_all_td == 'Connection Name' and str_all_td_1  == 'Wi-Fi\n':
#                         print (str_all_td_1)
                        all_td = all_tr_wifi[index-1].find_all(DB_TAG.TD)

                        print (all_td[4].text)
                        wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.MAC_ADDRESS).value = wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.MAC_ADDRESS).value + (all_td[4].text + '-' +str_all_td_1) + '\n'
                        wbk.save(str(out_xlsx_file))
                    if str_all_td == 'Connection Name' and str_all_td_1  == 'Ethernet\n':
                        print (str_all_td_1)
                        all_td = all_tr_wifi[index-1].find_all(DB_TAG.TD)
                        print (all_td[4].text)
                        wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.MAC_ADDRESS).value = wks.cell(row = DB_ROW.START_ROW, column = DB_COLUMN.MAC_ADDRESS).value + (all_td[4].text+ '-' + str_all_td_1) +'\n'
                        wbk.save(str(out_xlsx_file))
                index+=1
                    
        index_count += 1
    wbk.save(str(out_xlsx_file))
    wbk.close 
#     finally:
#         wbk.close   
    DB_ROW.START_ROW +=1    
    time2 = time.clock() - start_time
    print "Complete processing %s: %s" %(str(file), time2)       
#     print ('======================================================')
print ('Process is ended')
# name=raw_input('Enter your name : ')
# print ("Hi %s, Let us be friends!" % name)  
